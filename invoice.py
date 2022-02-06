import imp
from multiprocessing.context import ForkProcess
import openpyxl
from PIL import Image

wb = openpyxl.load_workbook("/home/jay/invoice_gen/data.xlsx")
sheet = wb["invoice2022"]
row_count = sheet.max_row

logo = Image.open('logo.png')
width, height = logo.size
ratio = width/height

company_mame = 'Scones & I.T.'
company_abn = 92230643882

def create_all_invoices():
    for i in range(2,row_count+1):
        invoice_number = sheet.cell(row = i, column =1).value
        customer = sheet.cell(row = i, column = 2).value
        abn = sheet.cell(row = i, column = 3).value
        amount = sheet.cell(row = i, column = 4).value
        description = sheet.cell(row = i, column = 5).value
        invoice_date = sheet.cell(row = i, column = 6).value
        due_date = sheet.cell(row = i, column = 7).value
        paid = sheet.cell(row = i, column = 8).value
        date_paid = sheet.cell(row = i, column = 9).value
        email = sheet.cell(row = i, column = 10).value
        
def create_invoice_by_abn(_abn):
    for i in range(2, row_count+1):
        invoice_number = sheet.cell(row = i, column =1).value
        abn = sheet.cell(row = i, column = 3).value
       
        if abn == _abn:
               print(invoice_number) 
        
def main():
    create_invoice_by_abn(57171900565)   
    
    
main()    
        

